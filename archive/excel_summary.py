#!/usr/bin/env python3
"""
Generate Excel summary sheets for batch grading
"""

import pandas as pd
import sqlite3
import os
from datetime import datetime
from typing import Dict, List, Any
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelSummaryGenerator:
    def __init__(self, output_dir: str = "grade_summaries"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_assignment_summary(self, grader, assignment_id: int, assignment_name: str) -> str:
        """Generate comprehensive Excel summary for an assignment"""
        
        # Get all submission data
        conn = sqlite3.connect(grader.db_path)
        
        # Get assignment info
        assignment_info = pd.read_sql_query("""
            SELECT * FROM assignments WHERE id = ?
        """, conn, params=(assignment_id,))
        
        # Get submissions with student info
        submissions = pd.read_sql_query("""
            SELECT s.*, st.name as student_name, st.email as student_email
            FROM submissions s
            LEFT JOIN students st ON s.student_id = st.student_id
            WHERE s.assignment_id = ?
            ORDER BY s.student_id
        """, conn, params=(assignment_id,))
        
        conn.close()
        
        if submissions.empty:
            return None
        
        # Analyze each submission for detailed breakdown
        detailed_data = self._analyze_all_submissions(submissions)
        
        # Create Excel file
        excel_path = self._create_excel_summary(
            assignment_name, assignment_info.iloc[0] if not assignment_info.empty else None,
            submissions, detailed_data
        )
        
        return excel_path
    
    def _analyze_all_submissions(self, submissions: pd.DataFrame) -> Dict[str, Dict]:
        """Analyze all submissions for detailed breakdown"""
        from detailed_analyzer import DetailedHomeworkAnalyzer
        
        analyzer = DetailedHomeworkAnalyzer()
        detailed_data = {}
        
        for _, submission in submissions.iterrows():
            student_id = submission['student_id']
            notebook_path = submission['notebook_path']
            
            if os.path.exists(notebook_path):
                try:
                    analysis = analyzer.analyze_notebook(notebook_path)
                    detailed_data[student_id] = analysis
                except Exception as e:
                    # Create error analysis
                    detailed_data[student_id] = {
                        'total_score': 0,
                        'element_scores': {},
                        'student_info': {'name': '', 'date': '', 'course': ''},
                        'error': str(e)
                    }
            else:
                detailed_data[student_id] = {
                    'total_score': 0,
                    'element_scores': {},
                    'student_info': {'name': '', 'date': '', 'course': ''},
                    'error': 'Notebook file not found'
                }
        
        return detailed_data
    
    def _create_excel_summary(self, assignment_name: str, assignment_info, submissions: pd.DataFrame, detailed_data: Dict) -> str:
        """Create the Excel summary file"""
        
        # Clean assignment name for filename
        safe_name = "".join(c for c in assignment_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        excel_filename = f"{safe_name}_Grade_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
        excel_path = os.path.join(self.output_dir, excel_filename)
        
        if OPENPYXL_AVAILABLE:
            return self._create_formatted_excel(excel_path, assignment_name, assignment_info, submissions, detailed_data)
        else:
            return self._create_basic_excel(excel_path, assignment_name, submissions, detailed_data)
    
    def _create_formatted_excel(self, excel_path: str, assignment_name: str, assignment_info, submissions: pd.DataFrame, detailed_data: Dict) -> str:
        """Create formatted Excel file with multiple sheets"""
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, assignment_name, assignment_info, submissions, detailed_data)
        
        # Create detailed breakdown sheet
        self._create_breakdown_sheet(wb, submissions, detailed_data)
        
        # Create individual feedback sheet
        self._create_feedback_sheet(wb, submissions, detailed_data)
        
        # Save workbook
        wb.save(excel_path)
        return excel_path
    
    def _create_summary_sheet(self, wb, assignment_name: str, assignment_info, submissions: pd.DataFrame, detailed_data: Dict):
        """Create the main summary sheet"""
        
        ws = wb.create_sheet("Grade Summary", 0)
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = f"Grade Summary: {assignment_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:H2')
        
        # Headers
        headers = [
            'Student ID', 'Student Name', 'Submission Date', 
            'Working Dir', 'Packages', 'Data Import', 'Data Inspection', 
            'Reflection Qs', 'Total Score', 'Percentage', 'Letter Grade'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        row = 5
        for _, submission in submissions.iterrows():
            student_id = submission['student_id']
            analysis = detailed_data.get(student_id, {})
            student_info = analysis.get('student_info', {})
            
            # Extract student name from notebook or use database name
            student_name = student_info.get('name', '') or submission.get('student_name', '') or student_id
            
            total_score = analysis.get('total_score', 0)
            percentage = (total_score / 37.5) * 100
            letter_grade = self._calculate_letter_grade(percentage)
            
            element_scores = analysis.get('element_scores', {})
            
            # Fill row data
            ws.cell(row=row, column=1, value=student_id)
            ws.cell(row=row, column=2, value=student_name)
            ws.cell(row=row, column=3, value=submission['submission_date'])
            ws.cell(row=row, column=4, value=element_scores.get('working_directory', 0))
            ws.cell(row=row, column=5, value=element_scores.get('package_loading', 0))
            ws.cell(row=row, column=6, value=element_scores.get('data_import', 0))
            ws.cell(row=row, column=7, value=element_scores.get('data_inspection', 0))
            ws.cell(row=row, column=8, value=element_scores.get('reflection_questions', 0))
            ws.cell(row=row, column=9, value=round(total_score, 1))
            ws.cell(row=row, column=10, value=f"{percentage:.1f}%")
            ws.cell(row=row, column=11, value=letter_grade)
            
            # Color code the letter grade
            grade_cell = ws.cell(row=row, column=11)
            if letter_grade in ['A', 'A-']:
                grade_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif letter_grade in ['B+', 'B', 'B-']:
                grade_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            elif letter_grade in ['C+', 'C', 'C-']:
                grade_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            elif letter_grade in ['D+', 'D', 'D-']:
                grade_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            else:
                grade_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row += 1
        
        # Add statistics
        stats_row = row + 2
        ws.cell(row=stats_row, column=1, value="Class Statistics:").font = Font(bold=True)
        
        scores = [detailed_data[sid].get('total_score', 0) for sid in detailed_data.keys()]
        if scores:
            ws.cell(row=stats_row + 1, column=1, value="Average:")
            ws.cell(row=stats_row + 1, column=2, value=f"{sum(scores)/len(scores):.1f}")
            ws.cell(row=stats_row + 2, column=1, value="Highest:")
            ws.cell(row=stats_row + 2, column=2, value=f"{max(scores):.1f}")
            ws.cell(row=stats_row + 3, column=1, value="Lowest:")
            ws.cell(row=stats_row + 3, column=2, value=f"{min(scores):.1f}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_breakdown_sheet(self, wb, submissions: pd.DataFrame, detailed_data: Dict):
        """Create detailed breakdown sheet"""
        
        ws = wb.create_sheet("Detailed Breakdown")
        
        # Headers
        headers = [
            'Student ID', 'Student Name', 'Working Directory (/2)', 'Package Loading (/4)', 
            'Data Import (/11)', 'Data Inspection (/8)', 'Reflection Questions (/12.5)',
            'Total (/37.5)', 'Percentage', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Data
        for row, (_, submission) in enumerate(submissions.iterrows(), 2):
            student_id = submission['student_id']
            analysis = detailed_data.get(student_id, {})
            student_info = analysis.get('student_info', {})
            element_scores = analysis.get('element_scores', {})
            
            student_name = student_info.get('name', '') or submission.get('student_name', '') or student_id
            total_score = analysis.get('total_score', 0)
            percentage = (total_score / 37.5) * 100
            
            # Notes about issues
            notes = []
            if analysis.get('missing_elements'):
                notes.append(f"Missing: {', '.join(analysis['missing_elements'][:2])}")
            if analysis.get('code_issues'):
                notes.append(f"Issues: {', '.join(analysis['code_issues'][:2])}")
            
            ws.cell(row=row, column=1, value=student_id)
            ws.cell(row=row, column=2, value=student_name)
            ws.cell(row=row, column=3, value=element_scores.get('working_directory', 0))
            ws.cell(row=row, column=4, value=element_scores.get('package_loading', 0))
            ws.cell(row=row, column=5, value=element_scores.get('data_import', 0))
            ws.cell(row=row, column=6, value=element_scores.get('data_inspection', 0))
            ws.cell(row=row, column=7, value=element_scores.get('reflection_questions', 0))
            ws.cell(row=row, column=8, value=round(total_score, 1))
            ws.cell(row=row, column=9, value=f"{percentage:.1f}%")
            ws.cell(row=row, column=10, value="; ".join(notes) if notes else "Complete")
    
    def _create_feedback_sheet(self, wb, submissions: pd.DataFrame, detailed_data: Dict):
        """Create sheet with condensed feedback for each student"""
        
        ws = wb.create_sheet("Student Feedback")
        
        # Headers
        ws.cell(row=1, column=1, value="Student Feedback Summary").font = Font(bold=True, size=14)
        
        row = 3
        for _, submission in submissions.iterrows():
            student_id = submission['student_id']
            analysis = detailed_data.get(student_id, {})
            student_info = analysis.get('student_info', {})
            
            student_name = student_info.get('name', '') or submission.get('student_name', '') or student_id
            total_score = analysis.get('total_score', 0)
            percentage = (total_score / 37.5) * 100
            
            # Student header
            ws.cell(row=row, column=1, value=f"{student_name} ({student_id})").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"Score: {total_score:.1f}/37.5 ({percentage:.1f}%)")
            
            row += 1
            
            # Key feedback points
            feedback_points = []
            
            element_scores = analysis.get('element_scores', {})
            if element_scores.get('working_directory', 0) < 1.5:
                feedback_points.append("• Working directory: Needs to run getwd() and show output")
            if element_scores.get('package_loading', 0) < 3:
                feedback_points.append("• Package loading: Issues with tidyverse/readxl")
            if element_scores.get('data_import', 0) < 8:
                feedback_points.append("• Data import: Missing or incomplete dataset imports")
            if element_scores.get('data_inspection', 0) < 6:
                feedback_points.append("• Data inspection: Need to run and show head(), str(), summary()")
            if element_scores.get('reflection_questions', 0) < 10:
                feedback_points.append("• Reflection questions: Need more detailed analytical responses")
            
            if not feedback_points:
                feedback_points.append("• Excellent work! All requirements completed successfully.")
            
            for point in feedback_points:
                ws.cell(row=row, column=1, value=point)
                row += 1
            
            row += 1  # Space between students
    
    def _create_basic_excel(self, excel_path: str, assignment_name: str, submissions: pd.DataFrame, detailed_data: Dict) -> str:
        """Create basic Excel file using pandas (fallback)"""
        
        # Prepare summary data
        summary_data = []
        
        for _, submission in submissions.iterrows():
            student_id = submission['student_id']
            analysis = detailed_data.get(student_id, {})
            student_info = analysis.get('student_info', {})
            element_scores = analysis.get('element_scores', {})
            
            student_name = student_info.get('name', '') or submission.get('student_name', '') or student_id
            total_score = analysis.get('total_score', 0)
            percentage = (total_score / 37.5) * 100
            
            summary_data.append({
                'Student_ID': student_id,
                'Student_Name': student_name,
                'Submission_Date': submission['submission_date'],
                'Working_Directory': element_scores.get('working_directory', 0),
                'Package_Loading': element_scores.get('package_loading', 0),
                'Data_Import': element_scores.get('data_import', 0),
                'Data_Inspection': element_scores.get('data_inspection', 0),
                'Reflection_Questions': element_scores.get('reflection_questions', 0),
                'Total_Score': round(total_score, 1),
                'Percentage': f"{percentage:.1f}%",
                'Letter_Grade': self._calculate_letter_grade(percentage),
                'AI_Score': submission.get('ai_score', ''),
                'Human_Score': submission.get('human_score', ''),
                'Final_Score': submission.get('final_score', '')
            })
        
        # Create DataFrame and save
        df = pd.DataFrame(summary_data)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Grade Summary', index=False)
            
            # Add class statistics
            stats_df = pd.DataFrame({
                'Statistic': ['Average', 'Highest', 'Lowest', 'Median', 'Students Submitted'],
                'Value': [
                    f"{df['Total_Score'].mean():.1f}",
                    f"{df['Total_Score'].max():.1f}",
                    f"{df['Total_Score'].min():.1f}",
                    f"{df['Total_Score'].median():.1f}",
                    len(df)
                ]
            })
            stats_df.to_excel(writer, sheet_name='Class Statistics', index=False)
        
        return excel_path
    
    def _calculate_letter_grade(self, percentage: float) -> str:
        """Calculate letter grade from percentage"""
        if percentage >= 97:
            return "A+"
        elif percentage >= 93:
            return "A"
        elif percentage >= 90:
            return "A-"
        elif percentage >= 87:
            return "B+"
        elif percentage >= 83:
            return "B"
        elif percentage >= 80:
            return "B-"
        elif percentage >= 77:
            return "C+"
        elif percentage >= 73:
            return "C"
        elif percentage >= 70:
            return "C-"
        elif percentage >= 67:
            return "D+"
        elif percentage >= 63:
            return "D"
        elif percentage >= 60:
            return "D-"
        else:
            return "F"

def generate_excel_summary_interface(grader, assignment_id: int, assignment_name: str):
    """Streamlit interface for generating Excel summary"""
    import streamlit as st
    
    st.subheader("📊 Generate Excel Grade Summary")
    
    with st.spinner("Analyzing all submissions and generating Excel summary..."):
        try:
            generator = ExcelSummaryGenerator()
            excel_path = generator.generate_assignment_summary(grader, assignment_id, assignment_name)
            
            if excel_path and os.path.exists(excel_path):
                st.success("✅ Excel summary generated successfully!")
                
                # Show file info
                file_size = os.path.getsize(excel_path) / 1024  # KB
                st.info(f"📁 File: {os.path.basename(excel_path)} ({file_size:.1f} KB)")
                
                # Download button
                with open(excel_path, 'rb') as f:
                    st.download_button(
                        label="📊 Download Excel Summary",
                        data=f.read(),
                        file_name=os.path.basename(excel_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # Show preview of data
                st.subheader("📋 Preview")
                
                # Read back the Excel file to show preview
                try:
                    preview_df = pd.read_excel(excel_path, sheet_name='Grade Summary')
                    st.dataframe(preview_df)
                except:
                    st.info("Excel file created successfully. Download to view contents.")
                
            else:
                st.error("Failed to generate Excel summary.")
                
        except Exception as e:
            st.error(f"Error generating Excel summary: {str(e)}")
            import traceback
            st.error(f"Details: {traceback.format_exc()}")